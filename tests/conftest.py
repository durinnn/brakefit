"""테스트용 KB 포맷 fixture 생성.

실제 거래내역을 아직 못 구했으므로, 실측으로 확인된 헤더 구조를 그대로 재현한
가짜 파일을 만들어 파서를 검증한다. 전부 지어낸 값이라 레포에 커밋해도 안전하다.

실 CSV 가 들어오면 이 fixture 는 버리지 말고 회귀 테스트로 남긴다.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

# 2026-08-27 실측 헤더 (kbsec.com 웹 거래내역 .xlsx)
LEDGER_HEADER = [
    "거래일", "내용", "종목명", "수량", "단가",
    "입금/입고/매도", "출금/출고/매수", "예수금잔액(원)",
]

# 값 표기는 아직 미확인이라 "있을 법한 것"들을 일부러 섞었다:
#   천단위 콤마 문자열 / 순수 숫자 / 매매 아닌 행 / 안내 행
LEDGER_ROWS = [
    ["2026-08-03", "입금",     "",         "",   "",       "1,000,000", "",          "1,000,000"],
    ["2026-08-04", "주식매수", "삼성전자", "10", "72,500", "",          "725,120",   "274,880"],
    ["2026-08-05", "주식매수", "삼성전자", "5",  "70,100", "",          "350,560",   "-75,680"],
    ["2026-08-11", "주식매도", "삼성전자", "8",  "74,300", "594,015",   "",          "518,335"],
    ["2026-08-12", "주식매수", "카카오",   "20", "41,250", "",          "825,140",   "-306,805"],
    ["2026-08-14", "배당금",   "삼성전자", "",   "",       "3,610",     "",          "-303,195"],
    ["2026-08-20", "주식매도", "카카오",   "20", "38,900", "776,880",   "",          "473,685"],
    ["합계",       "",         "",         "",   "",       "",          "",          ""],
]


@pytest.fixture
def ledger_xlsx(tmp_path: Path) -> Path:
    """거래내역(예수금 원장) 화면을 흉내낸 xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append(LEDGER_HEADER)
    for row in LEDGER_ROWS:
        ws.append(row)
    path = tmp_path / "kb_ledger_sample.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def ledger_with_preamble_xlsx(tmp_path: Path) -> Path:
    """HTS 쪽 export 처럼 헤더 위에 안내 행이 붙은 변형.

    헤더 행 번호를 하드코딩하면 여기서 깨진다 — 앵커 탐지가 필요한 이유.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["거래내역조회"])
    ws.append([])
    ws.append(["계좌번호:", "000-000-000-00", "", "", "", "", "출력일자:", "2026년08월27일"])
    ws.append(["계좌명:", "홍길동"])
    ws.append([])
    ws.append(LEDGER_HEADER)
    for row in LEDGER_ROWS:
        ws.append(row)
    path = tmp_path / "kb_ledger_preamble.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def empty_ledger_xlsx(tmp_path: Path) -> Path:
    """실제로 받은 빈 export — 헤더만 있고 데이터가 없는 경우."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append(LEDGER_HEADER)
    path = tmp_path / "kb_ledger_empty.xlsx"
    wb.save(path)
    return path


# ── 0330 일자별실현손익(상세) ────────────────────────────────────────────────
# 2026-08-27 화면 캡처에서 확인한 헤더
PNL_HEADER = [
    "매매일자", "종목명", "구분", "매매구분", "체결수량", "체결단가", "매수단가",
    "매수금액", "매도금액", "수수료", "제세금", "실현손익", "수익률",
]

PNL_ROWS = [
    ["2026-08-11", "삼성전자", "현금", "매도", "8",  "74,300", "72,500",
     "580,000", "594,400", "385", "1,308", "12,707", "2.19%"],
    ["2026-08-20", "카카오",   "현금", "매도", "20", "38,900", "41,250",
     "825,000", "778,000", "504",  "1,712", "(49,216)", "-5.97%"],
]


@pytest.fixture
def realized_pnl_xlsx(tmp_path: Path) -> Path:
    """0330 실현손익 화면을 흉내낸 xlsx (수수료·제세금 포함)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["일자별실현손익(상세)"])
    ws.append(["계좌번호:", "000-000-000-00", "", "", "", "", "출력일자:", "2026년08월27일"])
    ws.append([])
    ws.append(PNL_HEADER)
    for row in PNL_ROWS:
        ws.append(row)
    ws.append(["정상적으로 조회되었습니다."])
    path = tmp_path / "kb_0330_sample.xlsx"
    wb.save(path)
    return path


# ── 0112 거래내역조회 (2행 헤더 + 2행 레코드) ────────────────────────────────
# 2026-08-27 화면 캡처에서 확인한 구조. 한 거래가 두 줄에 걸쳐 있다.
TXN_HEADER_1 = ["거래일자", "거래종류", "수량", "거래금액", "정산금액",
                "거래세 등", "소득세", "양도세", "대출금", "유가잔고"]
TXN_HEADER_2 = ["", "종목명", "단가", "수수료", "펀드가입번호/신탁보수",
                "농특세/부가세", "지방소득세", "과세기준가", "신용/대출이자", "예수금"]


def _txn_pair(date, kind, name, qty, price, gross, settle, fee,
              tax_trade=0, tax_sur=0):
    """한 거래 = 두 줄."""
    return [
        [date, kind, qty, gross, settle, tax_trade, 0, 0, 0, ""],
        ["", name, price, fee, "", tax_sur, 0, "0.00", 0, ""],
    ]


@pytest.fixture
def transactions_xlsx(tmp_path: Path) -> Path:
    """0112 거래내역조회를 흉내낸 xlsx — 2행 헤더 + 2행 레코드."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["거래내역조회"])
    ws.append(["계좌번호:", "000-000-000-00"])
    ws.append([])
    ws.append(TXN_HEADER_1)
    ws.append(TXN_HEADER_2)

    # 실측한 입금 행 (매매 아님 → 스킵되어야 함)
    for line in _txn_pair("2026/08/27", "전자금융입금", "", "", "",
                          "10,000", "10,000", "0"):
        ws.append(line)
    # 매수
    for line in _txn_pair("2026/08/04", "현금매수", "삼성전자", "10", "72,500",
                          "725,000", "725,120", "120"):
        ws.append(line)
    # 매도 (거래세 + 농특세)
    for line in _txn_pair("2026/08/11", "현금매도", "삼성전자", "8", "74,300",
                          "594,400", "593,015", "98",
                          tax_trade="1,070", tax_sur="217"):
        ws.append(line)

    path = tmp_path / "kb_0112_sample.xlsx"
    wb.save(path)
    return path


# ── 0377 종목별주문/체결집계 ─────────────────────────────────────────────────
EXEC_HEADER = ["종목번호", "종목명", "매매구분", "주문수량", "체결수량",
               "취소수량", "미체결잔량", "체결금액", "체결단가"]


@pytest.fixture
def executions_xlsx(tmp_path: Path) -> Path:
    """0377 — 종목코드가 있는 유일한 화면."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(EXEC_HEADER)
    ws.append(["005930", "삼성전자", "매수", "10", "10", "0", "0", "725,000", "72,500"])
    ws.append([35720, "카카오", "매수", "20", "20", "0", "0", "825,000", "41,250"])
    path = tmp_path / "kb_0377_sample.xlsx"
    wb.save(path)
    return path
